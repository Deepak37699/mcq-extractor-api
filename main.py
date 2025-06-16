from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse
import re
import io
from typing import List, Dict, Any
import PyPDF2
from docx import Document
import openpyxl
from PIL import Image
import pytesseract
import cv2
import numpy as np
import fitz  # PyMuPDF
import os

# Configure Tesseract path for Windows
if os.name == 'nt':  # Windows
    tesseract_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    if os.path.exists(tesseract_path):
        pytesseract.pytesseract.tesseract_cmd = tesseract_path

app = FastAPI(
    title="MCQ Extractor API",
    description="API to extract Multiple Choice Questions from uploaded files",
    version="1.0.0"
)

@app.get("/")
async def root():
    return {"message": "MCQ Extractor API - Upload files to extract MCQ questions"}

@app.get("/health")
async def health():
    return {"status": "ok"}

class MCQExtractor:
    def __init__(self):
        # Simplified patterns for MCQ detection
        self.question_patterns = [
            r'^\d+\.\s+(.+)',  # 1. Question
            r'^Q\d+[:\.]?\s+(.+)',  # Q1: Question or Q1. Question
            r'^\d+\)\s+(.+)',  # 1) Question
            r'^Question:\s*\d+\s*(.+)',  # Question: 1 ... (PDF format)
            r'^\d+\s+(.+)',  # 1 Question (without dot/parenthesis)
            r'^(\d+)\s*[-\.]\s*(.+)',  # 1 - Question or 1. Question (flexible)
            r'^\d+[:\-\.]\s*(.+)',  # 1: Question, 1- Question, 1. Question (flexible)
        ]
        
        self.option_patterns = [
            r'^([A-Da-d])\)\s+(.+)',  # A) or a) Option
            r'^\(([A-Da-d])\)\s+(.+)',  # (A) or (a) Option
            r'^([A-Da-d])\.\s+(.+)',  # A. or a. Option
            r'^([A-Da-d])\s+(.+)',  # A Option or a Option (space separated)
        ]
        
        self.answer_patterns = [
            r'^Answer:\s*([A-Da-d])',
            r'^Ans:\s*([A-Da-d])',
            r'^Correct:\s*([A-Da-d])',
            r'^Answer\s*[:\-]\s*([A-Da-d])',
            r'^\s*([A-Da-d])\s*$',  # Just a letter on its own line
        ]

    def extract_text_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF file using multiple methods with page-by-page processing"""
        try:
            text = ""
            
            # Method 1: Try PyMuPDF first (better for complex layouts)
            try:
                pdf_document = fitz.open("pdf", file_content)
                for page_num in range(pdf_document.page_count):
                    page = pdf_document[page_num]
                    page_text = page.get_text()
                    if page_text.strip():
                        text += f"\n--- Page {page_num + 1} ---\n"
                        text += page_text + "\n"
                pdf_document.close()
                
                # If PyMuPDF gave us meaningful content, return it
                if text.strip() and len(text.strip()) > 100:
                    return text
            except Exception as e:
                print(f"PyMuPDF extraction failed: {e}")
            
            # Method 2: Try PyPDF2 as fallback
            try:
                pdf_file = io.BytesIO(file_content)
                pdf_reader = PyPDF2.PdfReader(pdf_file)
                text = ""
                
                for page_num, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    if page_text.strip():
                        text += f"\n--- Page {page_num + 1} ---\n"
                        text += page_text + "\n"
                
                if text.strip() and len(text.strip()) > 100:
                    return text
            except Exception as e:
                print(f"PyPDF2 extraction failed: {e}")
            
            # Method 3: If both fail, try OCR
            print("Text extraction failed, trying OCR...")
            return self._extract_text_from_scanned_pdf(file_content)
                
        except Exception as e:
            print(f"All PDF extraction methods failed: {e}")
            raise HTTPException(status_code=400, detail=f"Error reading PDF: {str(e)}")

    def _extract_text_from_scanned_pdf(self, file_content: bytes) -> str:
        """Extract text from scanned PDF using OCR"""
        try:
            # Check if tesseract is available
            import subprocess
            try:
                subprocess.run(['tesseract', '--version'], capture_output=True, check=True)
            except (subprocess.CalledProcessError, FileNotFoundError):
                return """Tesseract OCR is not installed or not in PATH. 

📋 To extract text from this scanned PDF, please install Tesseract OCR:

🔧 Windows Installation:
1. Download from: https://github.com/UB-Mannheim/tesseract/wiki
2. Run installer as Administrator  
3. Install to: C:\\Program Files\\Tesseract-OCR\\
4. ✅ Check "Add to PATH" during installation
5. Restart VS Code/Terminal

🧪 Test installation: tesseract --version

📖 See TESSERACT_INSTALL.md for detailed instructions.

🚀 Alternative: Convert PDF to text-based format using online tools."""

            print("Attempting OCR on scanned PDF...")
            pdf_document = fitz.open("pdf", file_content)
            text = ""
            
            for page_num in range(min(pdf_document.page_count, 5)):  # Limit to first 5 pages
                page = pdf_document[page_num]
                # Convert PDF page to image
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x zoom for better OCR
                img_data = pix.tobytes("png")
                
                # Use OCR to extract text from image
                image = Image.open(io.BytesIO(img_data))
                page_text = pytesseract.image_to_string(image, config='--psm 6')
                text += f"\n--- Page {page_num + 1} ---\n" + page_text + "\n"
            
            pdf_document.close()
            
            if text.strip():
                return text
            else:
                return "OCR completed but no readable text was found. The PDF may contain very low-quality images or non-text content."
                
        except Exception as e:
            return f"OCR extraction encountered an error: {str(e)}. Please check the TESSERACT_INSTALL.md file for setup instructions."

    def extract_text_from_docx(self, file_content: bytes) -> str:
        """Extract text from Word document"""
        try:
            doc_file = io.BytesIO(file_content)
            doc = Document(doc_file)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading Word document: {str(e)}")

    def extract_text_from_xlsx(self, file_content: bytes) -> str:
        """Extract text from Excel file"""
        try:
            excel_file = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(excel_file)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            return text
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")

    def extract_text_from_txt(self, file_content: bytes) -> str:
        """Extract text from plain text file"""
        try:
            return file_content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                return file_content.decode('latin-1')
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error reading text file: {str(e)}")

    def extract_text_from_image(self, file_content: bytes) -> str:
        """Extract text from image using OCR"""
        try:
            # Load image from bytes
            image = Image.open(io.BytesIO(file_content))
            
            # Convert PIL image to OpenCV format for preprocessing
            opencv_image = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
            
            # Preprocess image to improve OCR accuracy
            processed_image = self.preprocess_image(opencv_image)
            
            # Perform OCR
            text = pytesseract.image_to_string(processed_image, config='--psm 6')
            
            return text
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading image: {str(e)}")

    def preprocess_image(self, image):
        """Preprocess image to improve OCR accuracy"""
        try:
            # Convert to grayscale
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # Apply Gaussian blur to reduce noise
            blurred = cv2.GaussianBlur(gray, (5, 5), 0)
            
            # Apply adaptive thresholding to handle varying lighting
            thresh = cv2.adaptiveThreshold(
                blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2
            )
            
            # Apply morphological operations to clean up the image
            kernel = np.ones((1, 1), np.uint8)
            processed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
            processed = cv2.morphologyEx(processed, cv2.MORPH_OPEN, kernel)
            
            return processed
        except Exception as e:
            # If preprocessing fails, return original grayscale image
            return cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    def parse_mcqs(self, text: str) -> List[Dict[str, Any]]:
        """Parse MCQ questions from text with comprehensive logic"""
        # Preprocess text to handle common issues
        text = self._preprocess_text(text)
        
        mcqs = []
        lines = text.split('\n')
        
        # Clean and filter lines
        clean_lines = []
        for line in lines:
            line = line.strip()
            if line and len(line) > 1:  # Filter out very short lines
                clean_lines.append(line)
        
        # More aggressive approach - scan through all lines and look for question patterns
        i = 0
        while i < len(clean_lines):
            line = clean_lines[i]
            
            # Check for question patterns with more flexibility
            question_match = None
            question_number = None
            question_text = ""
            
            # Try each pattern
            for pattern in self.question_patterns:
                match = re.match(pattern, line)
                if match:
                    question_match = match
                    if len(match.groups()) == 1:
                        # Simple pattern like "1. Question text"
                        question_text = match.group(1).strip()
                        # Extract question number from the line
                        num_match = re.match(r'^(\d+)', line)
                        question_number = int(num_match.group(1)) if num_match else len(mcqs) + 1
                    elif len(match.groups()) == 2:
                        # Pattern with number and text like "1 - Question text"
                        question_number = int(match.group(1))
                        question_text = match.group(2).strip()
                    break
            
            if question_match and question_text:
                # Found a question - now collect options
                current_options = {}
                current_answer = None
                
                # Look ahead for options and question continuation
                j = i + 1
                
                # First, collect any continuation of the question text
                while j < len(clean_lines):
                    next_line = clean_lines[j]
                    next_line = self._clean_ocr_errors(next_line)
                    
                    # Check if this line contains options
                    if self._contains_multiple_options(next_line):
                        # Extract multiple options from this line
                        options = self._extract_multiple_options(next_line)
                        current_options.update(options)
                        j += 1
                        break
                    
                    # Check if this is a single option
                    is_single_option = False
                    for pattern in self.option_patterns:
                        option_match = re.match(pattern, next_line)
                        if option_match:
                            option_letter = option_match.group(1).upper()
                            option_text = option_match.group(2).strip()
                            current_options[option_letter] = option_text
                            is_single_option = True
                            break
                    
                    if is_single_option:
                        j += 1
                        break
                    
                    # Check if this is a new question
                    is_new_question = False
                    for pattern in self.question_patterns:
                        if re.match(pattern, next_line):
                            is_new_question = True
                            break
                    
                    if is_new_question:
                        break
                    
                    # Otherwise, this might be continuation of question text
                    if not current_options:  # Only add to question if we haven't found options yet
                        question_text += " " + next_line
                    
                    j += 1
                
                # Continue collecting options after the first option line/set
                while j < len(clean_lines) and len(current_options) < 4:
                    next_line = clean_lines[j]
                    next_line = self._clean_ocr_errors(next_line)
                    
                    # Try to extract more options
                    if self._contains_multiple_options(next_line):
                        options = self._extract_multiple_options(next_line)
                        current_options.update(options)
                        j += 1
                        continue
                    
                    # Check for single options
                    found_option = False
                    for pattern in self.option_patterns:
                        option_match = re.match(pattern, next_line)
                        if option_match:
                            option_letter = option_match.group(1).upper()
                            option_text = option_match.group(2).strip()
                            if option_letter not in current_options:
                                current_options[option_letter] = option_text
                            found_option = True
                            break
                    
                    if found_option:
                        j += 1
                        continue
                    
                    # Check for answers
                    answer_match = None
                    for pattern in self.answer_patterns:
                        match = re.match(pattern, next_line)
                        if match:
                            answer_match = match
                            current_answer = match.group(1).upper()
                            j += 1
                            break
                    
                    if answer_match:
                        continue
                    
                    # Check if we hit a new question
                    is_new_question = False
                    for pattern in self.question_patterns:
                        if re.match(pattern, next_line):
                            is_new_question = True
                            break
                    
                    if is_new_question:
                        break
                    
                    # Try to extract any remaining options from this line
                    remaining_options = self._extract_remaining_options(next_line, current_options)
                    if remaining_options:
                        current_options.update(remaining_options)
                    
                    j += 1
                
                # Save the question if it has at least one option
                if question_text and current_options:
                    # Validate and clean options
                    current_options = self._validate_options(current_options)
                    
                    mcqs.append({
                        "question_number": question_number if question_number else len(mcqs) + 1,
                        "question": question_text.strip(),
                        "options": current_options,
                        "correct_answer": current_answer
                    })
                
                # Continue from where we left off
                i = j
            else:
                i += 1
        
        # Sort questions by question number to maintain order
        mcqs.sort(key=lambda x: x['question_number'])
        
        return mcqs
    
    def _validate_options(self, options: dict) -> dict:
        """Validate and attempt to complete missing options"""
        # If we have less than 4 options, check if we can find missing ones
        if len(options) < 4:
            expected_letters = ['A', 'B', 'C', 'D']
            missing_letters = [letter for letter in expected_letters if letter not in options]
            
            # For debugging - we can log which options are missing
            if missing_letters and len(options) > 0:  # Only log if we have some options
                # Don't print during normal operation, but we could log this
                pass
        
        return options
    
    def _contains_multiple_options(self, line: str) -> bool:
        """Check if a line contains multiple options like 'a. opt1 b. opt2 c. opt3 d. opt4'"""
        # Clean OCR errors first
        clean_line = self._clean_ocr_errors(line)
        
        # Count how many option patterns we find
        option_count = len(re.findall(r'[a-dA-D]\.', clean_line))
        
        # Also check for patterns like "c. option d. option" even with OCR errors
        if option_count < 2:
            # Check for patterns that might be split or have OCR errors
            potential_options = re.findall(r'[a-dA-DGO0]\s*[\.\(]', clean_line)
            option_count = len(potential_options)
        
        return option_count >= 2
    
    def _extract_multiple_options(self, line: str) -> dict:
        """Extract multiple options from a single line with OCR error handling"""
        options = {}
        
        # Handle common OCR errors
        line = self._clean_ocr_errors(line)
        
        # Try multiple approaches to extract options
        
        # Approach 1: Standard splitting by option patterns
        parts = re.split(r'([a-dA-D]\.)', line)
        
        current_letter = None
        for part in parts:
            part = part.strip()
            if re.match(r'^[a-dA-D]\.$', part):
                # This is an option letter
                current_letter = part[0].upper()
            elif current_letter and part:
                # This is option text
                # Clean up the text - remove any trailing option letters and numbers
                clean_text = re.sub(r'\s+[a-dA-D][\.\(].*$', '', part).strip()
                clean_text = re.sub(r'\s+\d+$', '', clean_text).strip()  # Remove trailing numbers
                if clean_text:
                    options[current_letter] = clean_text
                current_letter = None
        
        # Approach 2: If we didn't get enough options, try regex matching
        if len(options) < 2:
            option_matches = re.findall(r'([a-dA-D])\.\s*([^a-dA-D\.]+?)(?=\s*[a-dA-D]\.|$)', line)
            for letter, text in option_matches:
                clean_text = text.strip()
                if clean_text:
                    options[letter.upper()] = clean_text
        
        return options
    
    def _clean_ocr_errors(self, line: str) -> str:
        """Clean common OCR errors in option lines"""
        # Common OCR errors for option letters
        line = line.replace('G.', 'c.').replace('G. (', 'c. ')
        line = line.replace('0.', 'c.').replace('O.', 'c.')
        line = line.replace('(36', 'c. 36').replace('G.(', 'c.')
        line = line.replace('G (', 'c. ').replace('G(', 'c.')
        
        # Additional OCR error patterns
        line = line.replace('I.', 'l.').replace('l.', '1.')  # OCR mixing I, l, 1
        line = line.replace('S.', '5.').replace('B.', '6.') if not re.search(r'[a-d]\.', line) else line
        
        # Fix spacing issues
        line = re.sub(r'([a-d]\.)\s*([A-Z])', r'\1 \2', line)  # Ensure space after option letter
        line = re.sub(r'(\d+)\.([A-Z])', r'\1. \2', line)  # Fix "1.Question" to "1. Question"
        
        return line.strip()
    
    def _extract_remaining_options(self, line: str, existing_options: dict) -> dict:
        """Try to extract any remaining options from a line"""
        new_options = {}
        
        # Clean the line first
        line = self._clean_ocr_errors(line)
        
        # Look for any option patterns
        option_matches = re.findall(r'([a-dA-D])\.\s*([^a-dA-D\.]+?)(?=\s*[a-dA-D]\.|$)', line)
        
        for letter, text in option_matches:
            letter = letter.upper()
            text = text.strip()
            
            # Only add if we don't already have this option
            if letter not in existing_options and text:
                new_options[letter] = text
        
        return new_options

    def _get_pdf_recommendation(self, results: dict) -> str:
        """Get recommendation based on PDF extraction results"""
        pypdf2_status = results.get("pypdf2", {}).get("status", "failed")
        pymupdf_status = results.get("pymupdf", {}).get("status", "failed")
        
        if pypdf2_status == "success" or pymupdf_status == "success":
            return "Text extraction successful. PDF should work with MCQ extraction."
        elif pypdf2_status == "empty" and pymupdf_status == "empty":
            return "No text found. This appears to be a scanned/image-based PDF. Try OCR extraction or convert to text-based PDF."
        else:
            return "Text extraction failed. PDF may be corrupted, protected, or have complex formatting."

    def _preprocess_text(self, text: str) -> str:
        """Preprocess text to handle common PDF extraction issues"""
        lines = text.split('\n')
        processed_lines = []
        
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
                
            # Clean OCR errors
            line = self._clean_ocr_errors(line)
            
            # Check if this line starts with a question number but is cut off
            if re.match(r'^\d+\.\s*$', line) or re.match(r'^\d+\)\s*$', line):
                # Question number on its own line, merge with next line
                if i + 1 < len(lines) and lines[i + 1].strip():
                    next_line = lines[i + 1].strip()
                    line = line + " " + next_line
                    # Skip the next line since we merged it
                    lines[i + 1] = ""
            
            # Handle incomplete option lines that might be split
            if re.match(r'^[a-dA-D]\.\s*$', line) and i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line and not re.match(r'^[a-dA-D]\.', next_line):
                    line = line + " " + next_line
                    lines[i + 1] = ""
            
            processed_lines.append(line)
        
        return '\n'.join(processed_lines)

extractor = MCQExtractor()

@app.post("/extract-mcq")
async def extract_mcq(file: UploadFile = File(...)):
    """
    Extract MCQ questions from uploaded file.
    Supports: PDF, DOCX, XLSX, TXT, and Image files (JPG, PNG, JPEG, BMP, TIFF)
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")
    
    # Check file extension
    file_extension = file.filename.lower().split('.')[-1]
    supported_formats = ['pdf', 'docx', 'xlsx', 'txt', 'jpg', 'jpeg', 'png', 'bmp', 'tiff']
    
    if file_extension not in supported_formats:
        raise HTTPException(
            status_code=400, 
            detail=f"Unsupported file format. Supported formats: {', '.join(supported_formats)}"
        )
    
    try:
        # Read file content
        file_content = await file.read()
        
        # Extract text based on file type
        if file_extension == 'pdf':
            text = extractor.extract_text_from_pdf(file_content)
            # If PDF extraction returned minimal text, it's likely scanned - try OCR
            if not text or len(text.strip()) < 50:
                print("PDF appears to be scanned, attempting OCR extraction...")
                try:
                    text = extractor._extract_text_from_scanned_pdf(file_content)
                except Exception as e:
                    text = f"OCR extraction failed: {str(e)}. This appears to be a scanned PDF. Please ensure Tesseract OCR is installed for image processing."
        elif file_extension == 'docx':
            text = extractor.extract_text_from_docx(file_content)
        elif file_extension == 'xlsx':
            text = extractor.extract_text_from_xlsx(file_content)
        elif file_extension == 'txt':
            text = extractor.extract_text_from_txt(file_content)
        elif file_extension in ['jpg', 'jpeg', 'png', 'bmp', 'tiff']:
            text = extractor.extract_text_from_image(file_content)
        
        # Extract MCQs from text
        mcqs = extractor.parse_mcqs(text)
        
        return JSONResponse(content={
            "filename": file.filename,
            "file_type": file_extension,
            "total_questions": len(mcqs),
            "extracted_text_preview": text[:500] + "..." if len(text) > 500 else text,
            "mcqs": mcqs
        })
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@app.get("/supported-formats")
async def get_supported_formats():
    """Get list of supported file formats"""
    return {
        "supported_formats": [
            {
                "format": "PDF",
                "extension": ".pdf",
                "description": "Portable Document Format (text-based and scanned with OCR)"
            },
            {
                "format": "Word Document",
                "extension": ".docx",
                "description": "Microsoft Word Document"
            },
            {
                "format": "Excel Spreadsheet",
                "extension": ".xlsx",
                "description": "Microsoft Excel Spreadsheet"
            },
            {
                "format": "Text File",
                "extension": ".txt",
                "description": "Plain Text File"
            },
            {
                "format": "JPEG Image",
                "extension": ".jpg/.jpeg",
                "description": "JPEG Image with OCR text extraction"
            },
            {
                "format": "PNG Image",
                "extension": ".png",
                "description": "PNG Image with OCR text extraction"
            },
            {
                "format": "BMP Image",
                "extension": ".bmp",
                "description": "BMP Image with OCR text extraction"
            },
            {
                "format": "TIFF Image",
                "extension": ".tiff",
                "description": "TIFF Image with OCR text extraction"
            }
        ],
        "special_endpoints": {
            "/extract-mcq": "Auto-detects file type and uses appropriate extraction method",
            "/extract-mcq-ocr": "Forces OCR extraction for scanned PDFs and images",
            "/debug-pdf": "Analyzes PDF extraction capabilities",
            "/test-ocr": "Tests OCR functionality on images"
        },
        "ocr_note": "OCR functionality requires Tesseract OCR to be installed on the system"
    }

@app.post("/test-ocr")
async def test_ocr(file: UploadFile = File(...)):
    """
    Test OCR functionality on an uploaded image.
    Returns extracted text without MCQ parsing.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")
    
    file_extension = file.filename.lower().split('.')[-1]
    
    if file_extension not in ['jpg', 'jpeg', 'png', 'bmp', 'tiff']:
        raise HTTPException(
            status_code=400, 
            detail="Only image files are supported for OCR testing"
        )
    
    try:
        file_content = await file.read()
        text = extractor.extract_text_from_image(file_content)
        
        return JSONResponse(content={
            "filename": file.filename,
            "extracted_text": text,
            "text_length": len(text),
            "confidence": "OCR completed successfully"
        })
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR failed: {str(e)}")

@app.post("/debug-pdf")
async def debug_pdf(file: UploadFile = File(...)):
    """
    Debug PDF extraction - shows what text extraction methods find.
    Useful for troubleshooting PDF parsing issues.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")
    
    file_extension = file.filename.lower().split('.')[-1]
    
    if file_extension != 'pdf':
        raise HTTPException(status_code=400, detail="Only PDF files are supported for PDF debugging")
    
    try:
        file_content = await file.read()
        results = {}
        
        # Try PyPDF2
        try:
            pdf_file = io.BytesIO(file_content)
            pdf_reader = PyPDF2.PdfReader(pdf_file)
            pypdf2_text = ""
            for page in pdf_reader.pages:
                pypdf2_text += page.extract_text() + "\n"
            results["pypdf2"] = {
                "text_length": len(pypdf2_text.strip()),
                "preview": pypdf2_text[:300] + "..." if len(pypdf2_text) > 300 else pypdf2_text,
                "status": "success" if pypdf2_text.strip() else "empty"
            }
        except Exception as e:
            results["pypdf2"] = {"status": "failed", "error": str(e)}
        
        # Try PyMuPDF
        try:
            pdf_document = fitz.open("pdf", file_content)
            pymupdf_text = ""
            for page_num in range(pdf_document.page_count):
                page = pdf_document[page_num]
                pymupdf_text += page.get_text() + "\n"
            pdf_document.close()
            results["pymupdf"] = {
                "text_length": len(pymupdf_text.strip()),
                "preview": pymupdf_text[:300] + "..." if len(pymupdf_text) > 300 else pymupdf_text,
                "status": "success" if pymupdf_text.strip() else "empty"
            }
        except Exception as e:
            results["pymupdf"] = {"status": "failed", "error": str(e)}
        
        # PDF info
        try:
            pdf_document = fitz.open("pdf", file_content)
            results["pdf_info"] = {
                "page_count": pdf_document.page_count,
                "is_encrypted": pdf_document.is_encrypted,
                "metadata": pdf_document.metadata
            }
            pdf_document.close()
        except Exception as e:
            results["pdf_info"] = {"error": str(e)}
        
        return JSONResponse(content={
            "filename": file.filename,
            "extraction_results": results,
            "recommendation": extractor._get_pdf_recommendation(results)
        })
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analyzing PDF: {str(e)}")

@app.post("/extract-mcq-ocr")
async def extract_mcq_with_ocr(file: UploadFile = File(...)):
    """
    Force OCR extraction for scanned PDFs and images.
    Use this when regular extraction fails on image-based documents.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")
    
    file_extension = file.filename.lower().split('.')[-1]
    supported_formats = ['pdf', 'jpg', 'jpeg', 'png', 'bmp', 'tiff']
    
    if file_extension not in supported_formats:
        raise HTTPException(
            status_code=400, 
            detail=f"OCR extraction supports: {', '.join(supported_formats)}"
        )
    
    try:
        file_content = await file.read()
        
        # Force OCR extraction
        if file_extension == 'pdf':
            text = extractor._extract_text_from_scanned_pdf(file_content)
        else:
            text = extractor.extract_text_from_image(file_content)
        
        # Check if OCR was successful
        if "OCR extraction failed" in text or "Tesseract" in text:
            return JSONResponse(content={
                "filename": file.filename,
                "file_type": file_extension,
                "status": "ocr_failed",
                "message": text,
                "solution": "Please install Tesseract OCR. See README.md for installation instructions."
            })
        
        # Parse MCQs from OCR text
        mcqs = extractor.parse_mcqs(text)
        
        return JSONResponse(content={
            "filename": file.filename,
            "file_type": file_extension,
            "extraction_method": "OCR",
            "total_questions": len(mcqs),
            "extracted_text_preview": text[:500] + "..." if len(text) > 500 else text,
            "mcqs": mcqs,
            "note": "This file was processed using OCR. Results may vary based on image quality."
        })
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"OCR processing failed: {str(e)}")

@app.get("/tesseract-status")
async def check_tesseract_status():
    """
    Check if Tesseract OCR is installed and working.
    Helpful for troubleshooting OCR issues.
    """
    try:
        import subprocess
        
        # Check if tesseract command exists
        try:
            result = subprocess.run(['tesseract', '--version'], 
                                 capture_output=True, text=True, check=True)
            version_info = result.stdout.strip()
            
            # Check available languages
            try:
                lang_result = subprocess.run(['tesseract', '--list-langs'], 
                                           capture_output=True, text=True, check=True)
                languages = lang_result.stdout.strip().split('\n')[1:]  # Skip first line
            except:
                languages = ["Unable to detect languages"]
                
            return JSONResponse(content={
                "tesseract_installed": True,
                "version": version_info.split('\n')[0] if version_info else "Unknown",
                "available_languages": languages[:10],  # Show first 10 languages
                "status": "✅ Tesseract is working correctly",
                "ocr_endpoints_ready": True
            })
            
        except (subprocess.CalledProcessError, FileNotFoundError) as e:
            return JSONResponse(content={
                "tesseract_installed": False,
                "error": str(e),
                "status": "❌ Tesseract not found in PATH",
                "installation_guide": "See TESSERACT_INSTALL.md for installation instructions",
                "download_url": "https://github.com/UB-Mannheim/tesseract/wiki",
                "ocr_endpoints_ready": False
            })
            
    except Exception as e:
        return JSONResponse(content={
            "error": f"Error checking Tesseract: {str(e)}",
            "status": "❌ Unable to check Tesseract status"
        })

@app.post("/extract-mcq-detailed")
async def extract_mcq_detailed(file: UploadFile = File(...)):
    """
    Extract MCQ questions with detailed extraction information.
    Useful for debugging and seeing what's happening during extraction.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file uploaded")
    
    file_extension = file.filename.lower().split('.')[-1]
    supported_formats = ['pdf', 'docx', 'xlsx', 'txt', 'jpg', 'jpeg', 'png', 'bmp', 'tiff']
    
    if file_extension not in supported_formats:
        raise HTTPException(
            status_code=400, 
            detail=f"Unsupported file format. Supported formats: {', '.join(supported_formats)}"
        )
    
    try:
        file_content = await file.read()
        
        # Extract text with method tracking
        extraction_method = "unknown"
        text = ""
        
        if file_extension == 'pdf':
            try:
                text = extractor.extract_text_from_pdf(file_content)
                extraction_method = "PDF text extraction"
                if not text.strip() or len(text.strip()) < 50:
                    text = extractor._extract_text_from_scanned_pdf(file_content)
                    extraction_method = "PDF OCR extraction"
            except Exception as e:
                text = extractor._extract_text_from_scanned_pdf(file_content)
                extraction_method = "PDF OCR extraction (fallback)"
        
        elif file_extension == 'docx':
            text = extractor.extract_text_from_docx(file_content)
            extraction_method = "DOCX text extraction"
        
        elif file_extension == 'xlsx':
            text = extractor.extract_text_from_xlsx(file_content)
            extraction_method = "XLSX text extraction"
        
        elif file_extension == 'txt':
            text = file_content.decode('utf-8')
            extraction_method = "Plain text"
        
        else:  # Image formats
            text = extractor.extract_text_from_image(file_content)
            extraction_method = "Image OCR extraction"
        
        if not text.strip():
            raise HTTPException(status_code=400, detail="No text could be extracted from the file")
        
        # Parse MCQs
        mcqs = extractor.parse_mcqs(text)
        
        # Analyze extraction quality
        lines = text.split('\n')
        non_empty_lines = [line.strip() for line in lines if line.strip()]
        
        # Count option completeness
        complete_questions = sum(1 for mcq in mcqs if len(mcq['options']) == 4)
        incomplete_questions = sum(1 for mcq in mcqs if 0 < len(mcq['options']) < 4)
        
        # Get statistics about options
        option_stats = {}
        for mcq in mcqs:
            option_count = len(mcq['options'])
            if option_count in option_stats:
                option_stats[option_count] += 1
            else:
                option_stats[option_count] = 1
        
        return JSONResponse(content={
            "filename": file.filename,
            "file_type": file_extension,
            "extraction_method": extraction_method,
            "total_questions": len(mcqs),
            "complete_questions": complete_questions,
            "incomplete_questions": incomplete_questions,
            "option_statistics": option_stats,
            "text_analysis": {
                "total_characters": len(text),
                "total_lines": len(lines),
                "non_empty_lines": len(non_empty_lines),
                "avg_line_length": sum(len(line) for line in non_empty_lines) / len(non_empty_lines) if non_empty_lines else 0
            },
            "extracted_text_preview": text[:1000] + "..." if len(text) > 1000 else text,
            "mcqs": mcqs
        })
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")
