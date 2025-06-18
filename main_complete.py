from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.responses import JSONResponse
import re
import json
import base64
import io
import os
from typing import List, Dict, Any, Optional, Union, Tuple
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
import numpy as np
import cv2
import sympy
from sympy import latex, sympify, parse_expr
import matplotlib.pyplot as plt
from tabulate import tabulate
import PyPDF2
from docx import Document
import openpyxl

# Configure Tesseract path for Windows
if os.name == 'nt':  # Windows
    tesseract_path = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
    if os.path.exists(tesseract_path):
        pytesseract.pytesseract.tesseract_cmd = tesseract_path

app = FastAPI(
    title="MCQ Extractor API",
    description="API to extract Multiple Choice Questions from uploaded files",
    version="1.0.0"
)

@app.get("/")
async def root():
    return {"message": "MCQ Extractor API - Upload files to extract MCQ questions"}

@app.get("/health")
async def health():
    return {"status": "ok"}

class MCQExtractor:
    def __init__(self):
        self.question_patterns = [
            r'^\d+\.\s+(.+)',  # 1. Question
            r'^Q\d+[:\.]?\s+(.+)',  # Q1: Question or Q1. Question
            r'^\d+\)\s+(.+)',  # 1) Question
            r'^Question:\s*\d+\s*(.+)',  # Question: 1 ... (PDF format)
            r'^\d+\s+(.+)',  # 1 Question (without dot/parenthesis)
            r'^(\d+)\s*[-\.]\s*(.+)',  # 1 - Question or 1. Question (flexible)
            r'^\d+[:\-\.]\s*(.+)',  # 1: Question, 1- Question, 1. Question (flexible)
            r'^Question\s+(\d+)[:\.]?\s*(.+)',  # Question 1: ... or Question 1. ...
            r'^(\d+)\s*[\)\.]?\s*(.+)',  # Very flexible: number followed by anything
            r'^\d+\s*[:\.]\s*(.+)',  # Number with colon or dot: 1: Question, 1. Question
        ]
        
        self.option_patterns = [
            r'^([A-Da-d])\)\s+(.+)',  # A) or a) Option
            r'^\(([A-Da-d])\)\s+(.+)',  # (A) or (a) Option
            r'^([A-Da-d])\.\s+(.+)',  # A. or a. Option
            r'^([A-Da-d])\s+(.+)',  # A Option or a Option (space separated)
            r'^([A-Da-d])\)\s*(.+)',  # A) Option (flexible spacing)
            r'^([A-Da-d])[\.\):\-]\s*(.+)',  # A. A) A: A- Option (very flexible)
            r'^\(([A-Da-d])\)\s*(.+)',  # (A) Option (flexible spacing)
            r'^([A-Da-d])\s*[\.\)]\s*(.+)',  # Flexible: A. or A) with flexible spacing
            # Enhanced patterns for incomplete options
            r'^([A-Da-d])[\.\)]\s*(Rs\.\s*\d+)',  # A. Rs. 1200
            r'^([A-Da-d])[\.\)]\s*(BS\.\s*\d+)',  # A. BS. 2020
            r'^([A-Da-d])[\.\)]\s*(\d+(?:\.\d+)?)',  # A. 123 or A. 12.5
            r'^([A-Da-d])[\.\)]\s*([A-Z][^A-D\n]*)',  # A. Capital letter start
        ]
        
        self.answer_patterns = [
            r'^Answer:\s*([A-Da-d])',
            r'^Ans:\s*([A-Da-d])',
            r'^Correct:\s*([A-Da-d])',
            r'^Answer\s*[:\-]?\s*([A-Da-d])',
            r'^Correct\s*Answer\s*[:\-]?\s*([A-Da-d])',
            r'^Ans\s*[:\-]?\s*([A-Da-d])',
            r'\b([A-Da-d])\s*\)\s*$',  # Answer at end of line
            r'^\s*([A-Da-d])\s*$',     # Single letter answer
            r'Answer\s*is\s*([A-Da-d])',
            r'Correct\s*option\s*[:\-]?\s*([A-Da-d])',
            r'\b([A-Da-d])\s*is\s*correct',
            r'Option\s*([A-Da-d])\s*is\s*correct',
        ]

    def extract_text_from_pdf(self, file_content: bytes) -> str:
        """Extract text from PDF file using multiple methods with page-by-page processing"""
        try:
            text = ""
            
            # Method 1: Try PyMuPDF (fitz) first - usually better for complex layouts
            try:
                pdf_doc = fitz.open(stream=file_content, filetype="pdf")
                for page_num in range(pdf_doc.page_count):
                    page = pdf_doc.page(page_num)
                    page_text = page.get_text()
                    if page_text.strip():
                        text += f"\n--- Page {page_num + 1} ---\n{page_text}\n"
                pdf_doc.close()
                
                if text.strip():
                    return self._preprocess_text(text)
            except Exception as e:
                print(f"PyMuPDF extraction failed: {e}")
            
            # Method 2: Try PyPDF2 as fallback
            try:
                pdf_reader = PyPDF2.PdfReader(io.BytesIO(file_content))
                for page_num, page in enumerate(pdf_reader.pages):
                    page_text = page.extract_text()
                    if page_text.strip():
                        text += f"\n--- Page {page_num + 1} ---\n{page_text}\n"
                
                if text.strip():
                    return self._preprocess_text(text)
            except Exception as e:
                print(f"PyPDF2 extraction failed: {e}")
            
            # Method 3: OCR as last resort for scanned PDFs
            try:
                return self._extract_text_from_scanned_pdf(file_content)
            except Exception as e:
                print(f"OCR extraction failed: {e}")
                raise HTTPException(status_code=400, detail="Could not extract text from PDF. It may be corrupted, protected, or contain only images.")
                
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading PDF: {str(e)}")

    def _extract_text_from_scanned_pdf(self, file_content: bytes) -> str:
        """Extract text from scanned PDF using OCR"""
        try:
            # Convert PDF to images using PyMuPDF
            pdf_doc = fitz.open(stream=file_content, filetype="pdf")
            text = ""
            
            for page_num in range(pdf_doc.page_count):
                # Get page as image
                page = pdf_doc.page(page_num)
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))  # 2x scaling for better OCR
                img_data = pix.tobytes("png")
                
                # Convert to PIL Image
                image = Image.open(io.BytesIO(img_data))
                
                # Convert to OpenCV format for preprocessing
                opencv_image = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
                
                # Preprocess image for better OCR
                processed_image = self.preprocess_image(opencv_image)
                
                # Perform OCR
                page_text = pytesseract.image_to_string(processed_image, config='--psm 6')
                
                if page_text.strip():
                    text += f"\n--- Page {page_num + 1} (OCR) ---\n{page_text}\n"
            
            pdf_doc.close()
            
            if not text.strip():
                raise HTTPException(status_code=400, detail="No text could be extracted from the scanned PDF")
            
            return self._preprocess_text(text)
            
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error processing scanned PDF: {str(e)}")

    def extract_text_from_docx(self, file_content: bytes) -> str:
        """Extract text from DOCX file"""
        try:
            doc = Document(io.BytesIO(file_content))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading DOCX file: {str(e)}")

    def extract_text_from_xlsx(self, file_content: bytes) -> str:
        """Extract text from Excel file"""
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(file_content))
            text = ""
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n--- Sheet: {sheet_name} ---\n"
                
                for row in sheet.iter_rows(values_only=True):
                    row_text = []
                    for cell in row:
                        if cell is not None:
                            row_text.append(str(cell))
                    if row_text:
                        text += " | ".join(row_text) + "\n"
            
            return text
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")

    def extract_text_from_txt(self, file_content: bytes) -> str:
        """Extract text from plain text file"""
        try:
            return file_content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                return file_content.decode('latin-1')
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Error reading text file: {str(e)}")

    def extract_text_from_image(self, file_content: bytes) -> str:
        """Extract text from image using OCR"""
        try:
            # Load image from bytes
            image = Image.open(io.BytesIO(file_content))
            
            # Convert PIL image to OpenCV format for preprocessing
            opencv_image = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2BGR)
            
            # Preprocess image to improve OCR accuracy
            processed_image = self.preprocess_image(opencv_image)
            
            # Perform OCR
            text = pytesseract.image_to_string(processed_image, config='--psm 6')
            
            return text
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Error reading image: {str(e)}")

    def preprocess_image(self, image):
        """Preprocess image to improve OCR accuracy"""
        try:
            # Convert to grayscale
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
            
            # Apply Gaussian blur to reduce noise
            blurred = cv2.GaussianBlur(gray, (5, 5), 0)
            
            # Apply adaptive thresholding to handle varying lighting
            thresh = cv2.adaptiveThreshold(
                blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2
            )
            
            # Apply morphological operations to clean up the image
            kernel = np.ones((1, 1), np.uint8)
            processed = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)
            processed = cv2.morphologyEx(processed, cv2.MORPH_OPEN, kernel)
            
            return processed
        except Exception as e:
            print(f"Error in image preprocessing: {e}")
            # Return original image if preprocessing fails
            return cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    def parse_mcqs(self, text: str) -> List[Dict[str, Any]]:
        """Parse MCQs from text with enhanced extraction"""
        mcqs = []
        
        # Preprocess the text
        text = self._preprocess_text(text)
        
        # Split into lines and clean
        lines = text.split('\n')
        clean_lines = []
        
        for line in lines:
            line = line.strip()
            if line:
                clean_lines.append(line)
        
        i = 0
        while i < len(clean_lines):
            line = clean_lines[i]
            
            # Skip empty lines
            if not line:
                i += 1
                continue
            
            # Clean OCR errors for this line
            line = self._clean_ocr_errors(line)
            
            # Try to match this line as a question
            question_match = None
            question_text = ""
            question_number = None
            
            for pattern in self.question_patterns:
                match = re.match(pattern, line)
                if match:
                    question_match = match
                    if len(match.groups()) == 1:
                        # Pattern without number group
                        question_text = match.group(1).strip()
                        # Extract question number from the line
                        num_match = re.match(r'^(\d+)', line)
                        question_number = int(num_match.group(1)) if num_match else len(mcqs) + 1
                    elif len(match.groups()) == 2:
                        # Pattern with number and text like "1 - Question text"
                        question_number = int(match.group(1))
                        question_text = match.group(2).strip()
                    break
            
            if question_match and question_text:
                # Found a question - now collect options using enhanced method
                current_options, next_i = self._extract_options_with_context(clean_lines, i + 1)
                
                # If we didn't get enough options, try the original method as fallback
                if len(current_options) < 2:
                    current_options = {}
                    current_answer = None
                    
                    # Fallback to original option extraction
                    j = i + 1
                    
                    # Collect options from following lines
                    while j < len(clean_lines) and len(current_options) < 4:
                        next_line = clean_lines[j]
                        next_line = self._clean_ocr_errors(next_line)
                        
                        # Check if this line contains options
                        if self._contains_multiple_options(next_line):
                            # Extract multiple options from this line
                            options = self._extract_multiple_options(next_line)
                            current_options.update(options)
                            j += 1
                            break
                        
                        # Check if this is a single option
                        is_single_option = False
                        for pattern in self.option_patterns:
                            option_match = re.match(pattern, next_line)
                            if option_match:
                                option_letter = option_match.group(1).upper()
                                option_text = option_match.group(2).strip()
                                current_options[option_letter] = option_text
                                is_single_option = True
                                break
                        
                        if is_single_option:
                            j += 1
                            continue
                        
                        # Check if this is a new question
                        is_new_question = False
                        for pattern in self.question_patterns:
                            if re.match(pattern, next_line):
                                is_new_question = True
                                break
                        
                        if is_new_question:
                            break
                        
                        j += 1
                    
                    i = j
                else:
                    i = next_i
                
                # Validate and create MCQ object
                if current_options:
                    current_options = self._validate_options(current_options)
                    
                    mcq = {
                        "question_number": question_number,
                        "question": question_text.strip(),
                        "options": current_options,
                        "correct_answer": None  # Will be filled by enhanced answer extraction
                    }
                    
                    mcqs.append(mcq)
            else:
                i += 1
        
        # Sort questions by question number to maintain order
        mcqs.sort(key=lambda x: x['question_number'])
        
        # Apply enhanced answer extraction
        mcqs = self._enhance_answer_extraction(text, mcqs)
        
        return mcqs

    def _preprocess_text(self, text: str) -> str:
        """Preprocess text to handle common PDF extraction issues"""
        lines = text.split('\n')
        processed_lines = []
        
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
                
            # Clean OCR errors
            line = self._clean_ocr_errors(line)
            
            # Check if this line starts with a question number but is cut off
            if re.match(r'^\d+\.\s*$', line) or re.match(r'^\d+\)\s*$', line):
                # Question number on its own line, merge with next line
                if i + 1 < len(lines) and lines[i + 1].strip():
                    next_line = lines[i + 1].strip()
                    line = line + " " + next_line
                    # Skip the next line since we merged it
                    lines[i + 1] = ""
            
            # Handle incomplete option lines that might be split
            if re.match(r'^[a-dA-D]\.\s*$', line) and i + 1 < len(lines):
                next_line = lines[i + 1].strip()
                if next_line and not re.match(r'^[a-dA-D]\.', next_line):
                    line = line + " " + next_line
                    lines[i + 1] = ""
            
            processed_lines.append(line)
        
        return '\n'.join(processed_lines)

    def _clean_ocr_errors(self, line: str) -> str:
        """Clean common OCR errors with improved validation"""
        # Store original line for corruption detection
        original = line
        
        # Check if this is likely a question line (starts with number followed by dot/space)
        if re.match(r'^\d+\.\s+', line):
            # This is likely a question - only fix minimal spacing issues
            line = re.sub(r'(\d+)\.([A-Z])', r'\1. \2', line)  # Fix "1.Question" to "1. Question"
            line = re.sub(r'(\d+)\.\s{2,}', r'\1. ', line)  # Fix multiple spaces after number
            return line.strip()
        
        # Check if this is likely an option line (starts with single letter followed by dot/parenthesis)
        is_option_line = re.match(r'^[a-dA-DGO0I]\s*[\.\(\)]', line)
        
        if is_option_line:
            # Apply OCR corrections only to option lines
            # Fix common OCR misrecognitions for option letters
            if re.match(r'^[GgO0][\.\(\)]', line):
                line = re.sub(r'^[Gg]([\.\(\)])', r'C\1', line)  # G -> C
                line = re.sub(r'^0([\.\(\)])', r'O\1', line)    # 0 -> O (but might be D)
            
            # Fix specific patterns
            line = line.replace('G.', 'C.').replace('G)', 'C)')
            line = line.replace('(G)', '(C)')
            
            # Fix spacing issues for options
            line = re.sub(r'([a-dA-D])\.([A-Z])', r'\1. \2', line)  # Ensure space after option letter
            line = re.sub(r'([a-dA-D])\)([A-Z])', r'\1) \2', line)  # Ensure space after option parenthesis
        
        # Validate that we haven't corrupted the text
        if self._is_corrupted_text(line) and not self._is_corrupted_text(original):
            # If cleaning caused corruption, return original
            return original.strip()
        
        return line.strip()

    def _is_corrupted_text(self, text: str) -> bool:
        """Detect if text appears to be corrupted by OCR or processing"""
        if not text:
            return False
        
        # Patterns that indicate corruption
        corruption_patterns = [
            r'[|}{]{2,}',                    # Multiple special characters
            r'[a-z]\s*[|]\s*[a-z]',         # Letters separated by pipes
            r'je\).*lelele',                 # Specific garbled pattern
            r'^[^A-Z0-9]*[a-z]{1,2}\s*[|]', # Starting with lowercase and pipes
            r'[a-z]{1}\s*[|]\s*[a-z]{1}',   # Single letters separated by pipes
            r'[}{|]{3,}',                   # Multiple braces/pipes
        ]
        
        for pattern in corruption_patterns:
            if re.search(pattern, text):
                return True
        
        # Check for extremely high ratio of special characters
        special_chars = sum(1 for c in text if c in '|}{[]()~`@#$%^&*+=<>')
        if len(text) > 0 and special_chars / len(text) > 0.3:
            return True
        
        return False

    def _contains_multiple_options(self, line: str) -> bool:
        """Check if a line contains multiple options"""
        option_count = len(re.findall(r'[a-dA-D]\.', line))
        return option_count >= 2

    def _extract_multiple_options(self, line: str) -> dict:
        """Extract multiple options from a single line with OCR error handling"""
        options = {}
        
        # Handle common OCR errors
        line = self._clean_ocr_errors(line)
        
        # Try multiple approaches to extract options
        
        # Approach 1: Standard splitting by option patterns
        parts = re.split(r'([a-dA-D]\.)', line)
        
        current_letter = None
        for part in parts:
            part = part.strip()
            if re.match(r'^[a-dA-D]\.$', part):
                # This is an option letter
                current_letter = part[0].upper()
            elif current_letter and part:
                # This is option text
                # Clean up the text - remove any trailing option letters and numbers
                clean_text = re.sub(r'\s+[a-dA-D][\.\(].*$', '', part).strip()
                clean_text = re.sub(r'\s+\d+$', '', clean_text).strip()  # Remove trailing numbers
                if clean_text:
                    options[current_letter] = clean_text
                current_letter = None
        
        # Approach 2: If we didn't get enough options, try regex matching
        if len(options) < 2:
            option_matches = re.findall(r'([a-dA-D])\.\s*([^a-dA-D\.]+?)(?=\s*[a-dA-D]\.|$)', line)
            for letter, text in option_matches:
                clean_text = text.strip()
                # Remove any trailing numbers or option letters
                clean_text = re.sub(r'\s+[a-dA-D][\.\(].*$', '', clean_text).strip()
                if clean_text:
                    options[letter.upper()] = clean_text
        
        return options

    def _validate_options(self, options: dict) -> dict:
        """Validate and attempt to complete missing options"""
        # If we have less than 4 options, check if we can find missing ones
        if len(options) < 4:
            # Check for common patterns in incomplete options
            for key, value in options.items():
                if value in ['Rs.', 'BS.'] or re.match(r'^(Rs\.|BS\.)\s*$', value):
                    # Mark as incomplete for potential fixing
                    options[key] = f"{value} [incomplete]"
        
        return options

    def _extract_options_with_context(self, lines: List[str], start_index: int) -> Tuple[Dict[str, str], int]:
        """Extract options with surrounding context to handle incomplete options"""
        options = {}
        current_index = start_index
        
        # Look for options in the next few lines
        max_lines_to_check = min(10, len(lines) - start_index)
        
        for i in range(max_lines_to_check):
            line_index = start_index + i
            if line_index >= len(lines):
                break
                
            line = lines[line_index].strip()
            if not line:
                continue
                
            # Clean the line
            clean_line = self._clean_ocr_errors(line)
            
            # Check if this is a new question (stop processing)
            for pattern in self.question_patterns:
                if re.match(pattern, clean_line):
                    if len(options) > 0:  # Only stop if we found some options
                        return options, line_index
            
            # Try to extract options from this line
            line_options = self._extract_options_from_single_line(clean_line, line_index, lines)
            
            if line_options:
                options.update(line_options)
                current_index = line_index + 1
                
                # If we have all 4 options, we're done
                if len(options) >= 4:
                    break
            else:
                # If no options found and we already have some, might be end of options
                if len(options) > 0:
                    break
        
        return options, current_index

    def _extract_options_from_single_line(self, line: str, line_index: int, all_lines: List[str]) -> Dict[str, str]:
        """Extract options from a single line with context awareness"""
        options = {}
        
        # Pattern to match incomplete options like "A. Rs." or "B. BS."
        incomplete_patterns = [
            r'^([A-Da-d])[\.\)]\s*(Rs\.|BS\.)\s*$',  # Just Rs. or BS.
            r'^([A-Da-d])[\.\)]\s*(Rs\.|BS\.)\s*(\d+)?',  # Rs./BS. with optional number
        ]
        
        # Check for incomplete options
        for pattern in incomplete_patterns:
            match = re.match(pattern, line)
            if match:
                letter = match.group(1).upper()
                prefix = match.group(2)
                number = match.group(3) if len(match.groups()) > 2 else None
                
                # Look for the complete value in surrounding lines
                complete_value = self._find_complete_option_value(line_index, all_lines, prefix, number)
                if complete_value:
                    options[letter] = complete_value
                else:
                    # Use what we have with a placeholder
                    options[letter] = f"{prefix} [value]"
                continue
        
        # Standard option extraction
        for pattern in self.option_patterns:
            match = re.match(pattern, line)
            if match:
                letter = match.group(1).upper()
                text = match.group(2).strip()
                
                # Skip if text is just incomplete markers
                if text not in ['Rs.', 'BS.', '']:
                    options[letter] = text
        
        return options

    def _find_complete_option_value(self, line_index: int, all_lines: List[str], prefix: str, partial_number: str = None) -> str:
        """Find complete option value by looking at surrounding context"""
        # Look in the next few lines for numbers that might complete the option
        search_range = min(3, len(all_lines) - line_index - 1)
        
        for i in range(1, search_range + 1):
            if line_index + i >= len(all_lines):
                break
                
            next_line = all_lines[line_index + i].strip()
            
            # Look for numbers that might be the missing value
            if prefix == "Rs.":
                # Look for currency amounts
                currency_match = re.search(r'(\d+(?:,\d{3})*(?:\.\d{2})?)', next_line)
                if currency_match:
                    return f"Rs. {currency_match.group(1)}"
            
            elif prefix == "BS.":
                # Look for year values
                year_match = re.search(r'(\d{4})', next_line)
                if year_match:
                    return f"BS. {year_match.group(1)}"
                    
                # Look for any 4-digit number
                number_match = re.search(r'(\d{4})', next_line)
                if number_match:
                    return f"BS. {number_match.group(1)}"
        
        # If partial number was provided, use it
        if partial_number:
            return f"{prefix} {partial_number}"
            
        return None

    def _enhance_answer_extraction(self, text: str, mcqs: List[Dict]) -> List[Dict]:
        """Enhanced answer extraction with multiple strategies"""
        enhanced_mcqs = []
        
        # Strategy 1: Look for answer key section
        answer_key = self._extract_answer_key(text)
        
        for i, mcq in enumerate(mcqs):
            enhanced_mcq = mcq.copy()
            
            # Try to find answer from answer key
            if answer_key and mcq.get('question_number'):
                q_num = mcq['question_number']
                if q_num in answer_key:
                    enhanced_mcq['correct_answer'] = answer_key[q_num]
                    enhanced_mcqs.append(enhanced_mcq)
                    continue
            
            # Strategy 2: Look for answer patterns near the question
            answer = self._find_answer_near_question(text, mcq)
            if answer:
                enhanced_mcq['correct_answer'] = answer
            
            enhanced_mcqs.append(enhanced_mcq)
        
        return enhanced_mcqs

    def _extract_answer_key(self, text: str) -> Dict[int, str]:
        """Extract answer key from text if present"""
        answer_key = {}
        
        # Look for answer key sections
        answer_key_patterns = [
            r'(?i)answer\s*key[:\s]*(.+?)(?=\n\n|\Z)',
            r'(?i)answers[:\s]*(.+?)(?=\n\n|\Z)',
            r'(?i)correct\s*answers[:\s]*(.+?)(?=\n\n|\Z)',
        ]
        
        for pattern in answer_key_patterns:
            match = re.search(pattern, text, re.DOTALL)
            if match:
                answer_section = match.group(1)
                
                # Extract individual answers
                answer_matches = re.findall(r'(\d+)[\.\):\s]*([A-Da-d])', answer_section)
                for q_num, answer in answer_matches:
                    answer_key[int(q_num)] = answer.upper()
        
        return answer_key

    def _find_answer_near_question(self, text: str, mcq: Dict) -> str:
        """Find answer near a specific question"""
        q_num = mcq.get('question_number')
        if not q_num:
            return None
        
        # Split text into lines
        lines = text.split('\n')
        
        # Find the question line
        question_line_index = None
        for i, line in enumerate(lines):
            if re.search(rf'\b{q_num}\.\s', line):
                question_line_index = i
                break
        
        if question_line_index is None:
            return None
        
        # Search in the next 10 lines for answer patterns
        search_end = min(question_line_index + 10, len(lines))
        
        for i in range(question_line_index, search_end):
            line = lines[i]
            
            for pattern in self.answer_patterns:
                match = re.search(pattern, line)
                if match:
                    return match.group(1).upper()
        
        return None

    def detect_math_content(self, text: str) -> Dict[str, Any]:
        """Detect mathematical content in the text"""
        math_indicators = {
            'has_math': False,
            'math_symbols': [],
            'equations': [],
            'formulas': [],
            'mathematical_expressions': [],
            'math_patterns': []
        }
        
        # Mathematical symbols and patterns
        math_symbols = [
            r'∫', r'∑', r'∏', r'√', r'∞', r'π', r'α', r'β', r'γ', r'δ', r'θ', r'λ', r'μ', r'σ', r'φ', r'ω',
            r'≤', r'≥', r'≠', r'≈', r'∈', r'∉', r'⊂', r'⊃', r'∪', r'∩', r'→', r'←', r'↔', r'∀', r'∃',
            r'sin', r'cos', r'tan', r'log', r'ln', r'exp', r'lim', r'sup', r'inf'
        ]
        
        # Equation patterns
        equation_patterns = [
            r'\d+\s*[+\-*/=]\s*\d+',  # Basic arithmetic
            r'[a-zA-Z]\s*[+\-*/=]\s*[a-zA-Z0-9]',  # Algebraic expressions
            r'\d+\^\d+',  # Exponents
            r'\d+/\d+',   # Fractions
            r'\(\s*\d+\s*[+\-*/]\s*\d+\s*\)',  # Expressions in parentheses
        ]
        
        # Check for math symbols
        for symbol in math_symbols:
            if re.search(symbol, text, re.IGNORECASE):
                math_indicators['math_symbols'].append(symbol)
                math_indicators['has_math'] = True
        
        # Check for equations
        for pattern in equation_patterns:
            matches = re.findall(pattern, text)
            if matches:
                math_indicators['equations'].extend(matches)
                math_indicators['has_math'] = True
        
        # Check for mathematical keywords
        math_keywords = ['calculate', 'solve', 'equation', 'formula', 'graph', 'function', 'derivative', 'integral', 'matrix']
        for keyword in math_keywords:
            if re.search(rf'\b{keyword}\b', text, re.IGNORECASE):
                math_indicators['math_patterns'].append(keyword)
                math_indicators['has_math'] = True
        
        return math_indicators

    def detect_visual_content(self, text: str) -> Dict[str, Any]:
        """Detect visual content references in text"""
        visual_indicators = {
            'has_visual_content': False,
            'table_references': [],
            'chart_references': [],
            'image_references': [],
            'diagram_references': [],
            'visual_patterns': [],
            'extracted_tables': []
        }
        
        # Visual content patterns
        table_patterns = [r'table\s+\d+', r'see\s+table', r'following\s+table', r'above\s+table']
        chart_patterns = [r'chart\s+\d+', r'graph\s+\d+', r'figure\s+\d+', r'diagram\s+\d+']
        image_patterns = [r'image\s+\d+', r'picture\s+\d+', r'photo\s+\d+']
        
        # Check for table references
        for pattern in table_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                visual_indicators['table_references'].extend(matches)
                visual_indicators['has_visual_content'] = True
        
        # Check for chart/graph references
        for pattern in chart_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                visual_indicators['chart_references'].extend(matches)
                visual_indicators['has_visual_content'] = True
        
        # Check for image references
        for pattern in image_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                visual_indicators['image_references'].extend(matches)
                visual_indicators['has_visual_content'] = True
        
        # Extract actual table structures
        tables = self._extract_tables(text)
        if tables:
            visual_indicators['extracted_tables'] = tables
            visual_indicators['has_visual_content'] = True
        
        return visual_indicators

    def _extract_tables(self, text: str) -> List[Dict]:
        """Extract table structures from text"""
        tables = []
        
        # Look for pipe-separated tables
        lines = text.split('\n')
        table_lines = []
        
        for line in lines:
            if '|' in line and line.count('|') >= 2:
                table_lines.append(line)
            elif table_lines and not line.strip():
                # End of table
                if len(table_lines) >= 2:
                    table = self._parse_table_lines(table_lines)
                    if table:
                        tables.append(table)
                table_lines = []
            elif table_lines:
                # Non-table line encountered, reset
                table_lines = []
        
        # Check for remaining table at end
        if len(table_lines) >= 2:
            table = self._parse_table_lines(table_lines)
            if table:
                tables.append(table)
        
        return tables

    def _parse_table_lines(self, lines: List[str]) -> Dict:
        """Parse lines into a table structure"""
        if not lines:
            return None
        
        rows = []
        for line in lines:
            # Split by pipe and clean
            cells = [cell.strip() for cell in line.split('|') if cell.strip()]
            if cells:
                rows.append(cells)
        
        if not rows:
            return None
        
        return {
            'rows': len(rows),
            'columns': len(rows[0]) if rows else 0,
            'data': rows[:5],  # Limit to first 5 rows for preview
            'preview': True if len(rows) > 5 else False
        }

    def _validate_and_fix_mcq(self, mcq: Dict) -> Dict:
        """Validate and fix common issues in MCQ"""
        fixed_mcq = mcq.copy()
        
        # Fix corrupted question text
        question = fixed_mcq.get('question', '')
        if self._is_corrupted_text(question):
            # Try to repair or mark as needs review
            fixed_mcq['question'] = f"[NEEDS REVIEW] {question}"
            fixed_mcq['extraction_issues'] = fixed_mcq.get('extraction_issues', []) + ['corrupted_question']
        
        # Fix incomplete options
        options = fixed_mcq.get('options', {})
        incomplete_options = []
        
        for opt_key, opt_value in options.items():
            if opt_value in ['Rs.', 'BS.', ''] or re.match(r'^(Rs\.|BS\.)\s*$', str(opt_value)):
                incomplete_options.append(opt_key)
        
        if incomplete_options:
            fixed_mcq['extraction_issues'] = fixed_mcq.get('extraction_issues', []) + ['incomplete_options']
        
        # Validate option completeness
        expected_options = {'A', 'B', 'C', 'D'}
        actual_options = set(options.keys())
        missing_options = expected_options - actual_options
        
        if missing_options:
            fixed_mcq['extraction_issues'] = fixed_mcq.get('extraction_issues', []) + ['missing_options']
            fixed_mcq['missing_options'] = list(missing_options)
        
        return fixed_mcq


# Initialize the MCQ extractor after class definition
mcq_extractor = MCQExtractor()

@app.post("/extract-mcq")
async def extract_mcqs(file: UploadFile = File(...)):
    """Extract MCQs from uploaded file with basic processing"""
    try:
        # Read file content
        content = await file.read()
        
        # Extract text based on file type
        file_extension = file.filename.split('.')[-1].lower()
        
        if file_extension == 'pdf':
            text = mcq_extractor.extract_text_from_pdf(content)
        elif file_extension in ['jpg', 'jpeg', 'png', 'bmp', 'tiff']:
            text = mcq_extractor.extract_text_from_image(content)
        elif file_extension == 'docx':
            text = mcq_extractor.extract_text_from_docx(content)
        elif file_extension in ['xlsx', 'xls']:
            text = mcq_extractor.extract_text_from_xlsx(content)
        elif file_extension == 'txt':
            text = mcq_extractor.extract_text_from_txt(content)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_extension}")
        
        if not text.strip():
            raise HTTPException(status_code=400, detail="No text could be extracted from the file")
        
        # Parse MCQs
        mcqs = mcq_extractor.parse_mcqs(text)
        
        if not mcqs:
            raise HTTPException(status_code=400, detail="No MCQs found in the text")
        
        # Count statistics
        total_questions = len(mcqs)
        complete_questions = sum(1 for mcq in mcqs if len(mcq.get('options', {})) >= 3)
        questions_with_answers = sum(1 for mcq in mcqs if mcq.get('correct_answer'))
        
        return {
            "success": True,
            "file_info": {
                "filename": file.filename,
                "file_type": file_extension.upper(),
                "file_size_mb": round(len(content) / (1024 * 1024), 2)
            },
            "extraction_summary": {
                "total_questions": total_questions,
                "complete_questions": complete_questions,
                "questions_with_answers": questions_with_answers
            },
            "mcqs": mcqs
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

@app.post("/extract-mcq-enhanced")
async def extract_mcqs_enhanced(file: UploadFile = File(...)):
    """Extract MCQs with enhanced processing, math detection, and visual content analysis"""
    try:
        # Read file content
        content = await file.read()
        
        # Extract text based on file type
        file_extension = file.filename.split('.')[-1].lower()
        
        if file_extension == 'pdf':
            text = mcq_extractor.extract_text_from_pdf(content)
        elif file_extension in ['jpg', 'jpeg', 'png', 'bmp', 'tiff']:
            text = mcq_extractor.extract_text_from_image(content)
        elif file_extension == 'docx':
            text = mcq_extractor.extract_text_from_docx(content)
        elif file_extension in ['xlsx', 'xls']:
            text = mcq_extractor.extract_text_from_xlsx(content)
        elif file_extension == 'txt':
            text = mcq_extractor.extract_text_from_txt(content)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_extension}")
        
        if not text.strip():
            raise HTTPException(status_code=400, detail="No text could be extracted from the file")
        
        # Parse MCQs with enhanced processing
        mcqs = mcq_extractor.parse_mcqs(text)
        
        if not mcqs:
            return {
                "success": False,
                "message": "No MCQs found in the text",
                "file_info": {
                    "filename": file.filename,
                    "file_type": file_extension.upper(),
                    "file_size_mb": round(len(content) / (1024 * 1024), 2)
                },
                "mcqs": []
            }
        
        # Apply enhanced analysis
        enhanced_mcqs = []
        for mcq in mcqs:
            enhanced_mcq = mcq_extractor._validate_and_fix_mcq(mcq)
            
            # Add content analysis
            question_text = enhanced_mcq.get('question', '')
            options_text = ' '.join(enhanced_mcq.get('options', {}).values())
            full_text = f"{question_text} {options_text}"
            
            # Detect math content
            math_analysis = mcq_extractor.detect_math_content(full_text)
            enhanced_mcq['content_analysis'] = {
                'mathematics': math_analysis
            }
            enhanced_mcq['math_content'] = math_analysis
            enhanced_mcq['has_math_content'] = math_analysis['has_math']
            
            # Detect visual content
            visual_analysis = mcq_extractor.detect_visual_content(full_text)
            enhanced_mcq['content_analysis']['visual_elements'] = visual_analysis
            enhanced_mcq['visual_content'] = visual_analysis
            enhanced_mcq['has_visual_content'] = visual_analysis['has_visual_content']
            
            # Classify question type
            if math_analysis['has_math']:
                enhanced_mcq['question_type'] = 'mathematical'
            elif visual_analysis['has_visual_content']:
                enhanced_mcq['question_type'] = 'visual'
            else:
                enhanced_mcq['question_type'] = 'standard'
            
            enhanced_mcqs.append(enhanced_mcq)
        
        # Document-level analysis
        doc_math_analysis = mcq_extractor.detect_math_content(text)
        doc_visual_analysis = mcq_extractor.detect_visual_content(text)
        
        # Count statistics
        total_questions = len(enhanced_mcqs)
        complete_questions = sum(1 for mcq in enhanced_mcqs if len(mcq.get('options', {})) >= 3)
        questions_with_answers = sum(1 for mcq in enhanced_mcqs if mcq.get('correct_answer'))
        mathematical_questions = sum(1 for mcq in enhanced_mcqs if mcq.get('has_math_content', False))
        visual_questions = sum(1 for mcq in enhanced_mcqs if mcq.get('has_visual_content', False))
        
        # Type distribution
        type_distribution = {}
        for mcq in enhanced_mcqs:
            q_type = mcq.get('question_type', 'standard')
            type_distribution[q_type] = type_distribution.get(q_type, 0) + 1
        
        # Processing notes
        processing_notes = []
        processing_notes.append(f"✅ Successfully extracted {total_questions} questions")
        if mathematical_questions > 0:
            processing_notes.append(f"🔢 Found {mathematical_questions} questions with mathematical content")
        if visual_questions > 0:
            processing_notes.append(f"📊 Found {visual_questions} questions with visual content")
        
        # Check for issues
        incomplete_questions = sum(1 for mcq in enhanced_mcqs if len(mcq.get('options', {})) < 4)
        if incomplete_questions > 0:
            processing_notes.append(f"⚠️ {incomplete_questions} questions have incomplete options")
        
        missing_answers = total_questions - questions_with_answers
        if missing_answers > 0:
            processing_notes.append(f"❓ {missing_answers} questions are missing answers")
        
        return {
            "success": True,
            "file_info": {
                "filename": file.filename,
                "file_type": file_extension.upper(),
                "file_size_mb": round(len(content) / (1024 * 1024), 2)
            },
            "extraction_summary": {
                "total_questions": total_questions,
                "mathematical_questions": mathematical_questions,
                "visual_content_questions": visual_questions,
                "complete_questions": complete_questions,
                "questions_with_answers": questions_with_answers,
                "question_type_distribution": type_distribution
            },
            "document_analysis": {
                "has_mathematical_content": doc_math_analysis['has_math'],
                "has_visual_content": doc_visual_analysis['has_visual_content'],
                "mathematical_elements": {
                    "symbols_found": len(doc_math_analysis['math_symbols']),
                    "equations_found": len(doc_math_analysis['equations']),
                    "formulas_found": len(doc_math_analysis['formulas']),
                    "math_patterns": doc_math_analysis['math_patterns']
                },
                "visual_elements": {
                    "table_references": len(doc_visual_analysis['table_references']),
                    "chart_references": len(doc_visual_analysis['chart_references']),
                    "image_references": len(doc_visual_analysis['image_references']),
                    "extracted_tables": len(doc_visual_analysis['extracted_tables'])
                }
            },
            "mcqs": enhanced_mcqs,
            "processing_notes": processing_notes,
            "enhanced_features": {
                "mathematical_notation_support": True,
                "visual_content_detection": True,
                "table_extraction": True,
                "equation_parsing": True,
                "content_type_classification": True
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
